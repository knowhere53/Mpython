# 1. 새창이 크롬에 뜨면, 로그인 후에 원하는 마이페이지 리스트를 첫창에!!! 띄워야 함.
from selenium import webdriver
# from bs4 import BeautifulSoup4
import re
import os
import time
import openpyxl
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
# 파이썬 폴더가 아니면 작업 폴더 바꿔주기

os.getcwd()  # 작업 폴더 확인
# os.chdir('C:\\Dropbox\\0. To Do\\0. 2년차 리뷰\\2. 네트워크 데이터') #작업폴더 변환. 이때 \\ 두번이다. 주의!
# os.getcwd()  #작업 폴더 재확인


"""이 시트는 마이폴더에 저장된 것중에서  -기업이름 + 주요 물품 + 기업현황/사업현황/사업구성의 구성을 갖고 오는 프로그램이다."""

""" 웹 크롤링 """
driver = webdriver.Chrome("/Users/yongjinjang/SynologyDrive/M파이썬/chromdriver/chromedriver")


# 웹 자원 로드를 위해 암묵적으로 딜레이
delay_time = 10
driver.implicitly_wait(delay_time)

# URL 접근
driver.get('https://www.kisline.com/cm/CM0100M00GE00.nice')

# ID, PW 입력
id = "knowhere53"
pw = "jyj8024!A"
driver.find_element(By.XPATH, '//*[@id="lgnuid"]').send_keys(id)
driver.find_element(By.XPATH, '//*[@id="lgnupassword"]').send_keys(pw)


# 로그인 버튼 클릭
driver.find_element(By.XPATH, '//*[@id="loginForm"]/div/div/fieldset/a').click()
wait = WebDriverWait(driver, 2)

filename = str()
# cited_list_filename = input("cited_list 파일명을 입력하시오:")
filename = "8.2`.KISLINE"
wb = openpyxl.Workbook()
wb.save(filename + '.xlsx')
wb = openpyxl.Workbook()
sheet = wb.active

driver.get('https://www.kisline.com/my/MY0100M00GE00.nice')  # 마이폴더 링크 이동
wait = WebDriverWait(driver, 1)

driver.find_element(By.XPATH, '//*[@id="grpseq"]').click()  # 선택박스
driver.find_element(By.XPATH, '//*[@id="grpseq"]/option[5]').click() #5번째, 1우주기업 선택
driver.find_element(By.XPATH, '//*[@id="myCcnEprSeach"]/span').click()  # 조회 버튼
wait = WebDriverWait(driver, 1)


a = driver.find_element(By.XPATH, '//*[@id="myCcnEprview"]/div[3]/div[3]').text  # 총페이지 수 있는 곳
page = int(a[a.rfind("/")+1 : a.rfind(" Page")])  # 총페이지 수
print(page,"page")
count = re.findall("\d+",(driver.find_element(By.XPATH, '//*[@id="myCcnEprview"]/p').text)) #나머지 건수. re.findall()은 숫자찾기
aa,count = divmod(int(count[0]),20)  #20으로 나눈 나머지는 count.
print(page,"page", " 마지막 페이지",count,"개.")

#page =2
#count=1
p=1
j=1
driver.get('https://www.kisline.com/my/MY0100M00GE00.nice')  # 마이폴더 링크 이동
driver.find_element(By.XPATH, '//*[@id="grpseq"]').click()  # 선택박스
driver.find_element(By.XPATH, '//*[@id="grpseq"]/option[5]').click()  # 5번째, 1우주기업 선택
driver.find_element(By.XPATH, '//*[@id="myCcnEprSeach"]/span').click()  # 조회 버튼
wait = WebDriverWait(driver, 1)
for p in range(1,page+1):
    if p == page: #마지막 페이지는 pp가 20으로 나눈 나머지만큼만.
        pp = count+1
    else:
        pp = 21
    for j in range(1, pp):
        driver.get('https://www.kisline.com/my/MY0100M00GE00.nice')  # 마이폴더 링크 이동
        driver.find_element(By.XPATH, '//*[@id="grpseq"]').click()  # 선택박스
        driver.find_element(By.XPATH, '//*[@id="grpseq"]/option[5]').click()  # 5번째, 1우주기업 선택
        driver.find_element(By.XPATH, '//*[@id="myCcnEprSeach"]/span').click()  # 조회 버튼
        time.sleep(2)
        if p>1:
            driver.find_element(By.XPATH, '//*[@id="myCcnEprview"]/div[3]/div[3]/a['+str(p+1)+']').click()  # 2page 이동
            time.sleep(1)  # sec. 무조건 1초 기다려라. 0.8 같이 소수점 가능. 무조건 됨. 잘됨. 굿! import time 필요.
        print(p,"페이지",j, "번째")
        driver.find_element(By.XPATH,
            '//*[@id="myCcnEprview"]/div[2]/table/tbody/tr[' + str(j) + ']/td[4]/a').click()  # 기업링크 이동
        delay_time = 10
        driver.implicitly_wait(delay_time)
        list = []
        firm = driver.find_element(By.XPATH, '//*[@id="overviewFrm"]/div[1]/dl/dt').text  # 기업명
        list.append(firm[:firm.rfind("(")])
        input_1 = driver.find_element(By.XPATH, '//*[@id="overviewFrm"]/div[3]/div[1]/div/table/tbody/tr[11]/td').text  # 주요상품
        list.append(input_1)
        print(list)
        driver.find_element(By.XPATH, '//*[@id="content"]/div/ul[4]/li/a').click()  # 기업재무 버튼
        driver.find_element(By.XPATH, '//*[@id="content"]/div/ul[4]/li/ul/li[1]/a').click()  # 개별재무 버튼
        driver.find_element(By.XPATH, '//*[@id="cont"]/ul/li[10]/a/span/span').click()  # my재무계정 탭
        time.sleep(1)  # sec. 무조건 1초 기다려라. 0.8 같이 소수점 가능. 무조건 됨. 잘됨. 굿! import time 필요.
        try:
            a = driver.find_element(By.CLASS_NAME,"cont")
            b = a.text.split("\n")
            list.extend([s for s in b if "개발비" in s][0].split(" "))
            list.extend([s for s in b if "개발비" in s][1].split(" "))
            list.extend([s for s in b if "개발비" in s][2].split(" "))
            list.extend([s for s in b if "자산총계" in s][0].split(" "))
            list.extend([s for s in b if "총매출액" in s][0].split(" "))
            list.extend([s for s in b if "당기순이익" in s][0].split(" "))
            list
            for l in list[2:]:
                if "," in l:
                    lr = l.replace(",", "")
                    lr = int(lr)
                    list[list.index(l)] = lr
#            for z in range(1, 5):
#                name = driver.find_element(By.XPATH, '//*[@id="tableFix"]/tbody/tr['+str(z)+']/td['+str(1)+']').text  # 기업명
#                list.append(name)
#                num = driver.find_elementBy.XPATH, '//*[@id="tableFix"]/tbody/tr['+str(z)+']/td['+str(2)+']').text  # 기업명
        except:
            h = 0
        print(list)
        sheet.append(list)

wb.save(filename + '.xlsx')
print('여기 저장 됨', filename + '.xlsx')
print("Done")


